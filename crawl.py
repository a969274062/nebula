import xlwt
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
import os
from idlelib.iomenu import encoding
import concurrent.futures
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionBuilder
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import subprocess
from selenium.webdriver.common.keys import Keys
from urllib.request import urlretrieve
import random
import requests
import json
import io
import base64
import numpy as np
from PIL import Image
import re
file_path = "论文信息表_3.xlsx"
# 定义列标题
columns = [
    "期刊名称", "论文标题", "作者姓名", "单位", "摘要", "关键词","基金资助",
    "发表时间","专辑","专题","分类号", "下载量", "页数", "引用量", "参考文献","正文"
]


def click_cancle(dr,xpath):
    try:
        # 点击取消
        cancel = dr.find_element(By.XPATH, xpath)
        dr.execute_script("arguments[0].click();", cancel)

    except Exception as e:
        print(e)

def cancle(driver):
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[2]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[3]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[4]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[7]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[8]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[9]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[10]/i')


from selenium import webdriver


def create_edge_driver(*, headless=False):
    # 配置浏览器选项
    options = webdriver.EdgeOptions()

    # 启用无头模式（可选）
    if headless:
        options.add_argument('--headless')


    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    # 初始化浏览器驱动
    browser = webdriver.Edge(options=options)
    #print(browser.title)

    return browser

def get_info(driver, xpath):
    try:
        element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return element.text
    except:
        return '无'

def crawl(start,key, page):
    # browser_path = "C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    # subprocess.Popen([browser_path, '--remote-debugging-port=9222'])
    driver = create_edge_driver(headless=False)
    time.sleep(1)
    driver.get("https://www.cnki.net/")  # 打开网页
    #driver.maximize_window()                      #最大化窗口
    time.sleep(1)  # 加载等待
    # txt_SearchText 为搜索框的位置  键入搜索内容
    driver.find_element(By.ID,'txt_SearchText').send_keys(key)
    # 只筛选期刊
    cancle(driver)
    # 定位到搜索按钮的位置并点击
    action = ActionChains(driver)
    action.move_to_element(driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[1]/div/div[1]/i')).perform()
    search = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[1]/div/div[2]/ul[3]/li[2]/a')
    driver.execute_script("arguments[0].click();", search)
    time.sleep(2)
    ActionBuilder(driver).clear_actions()
    # 定位到搜索按钮的位置并点击
    search = driver.find_element(By.XPATH,'/html/body/div[2]/div[2]/div[2]/div[1]/input[2]')
    driver.execute_script("arguments[0].click();", search)
    time.sleep(2) #等待加载
    # 获取初始所有窗口
    windows_init = driver.window_handles
    # 切换到目录页面对应的窗口
    driver.switch_to.window(windows_init[0])
    list_handle = driver.current_window_handle
    print(driver.title)
    try:
        for k in range(start,page):
            # page = 'page' + str(k)
            # driver.find_element(By.ID,page).click()
            if k == start:
                for t in range(1, start):
                    ActionChains(driver).send_keys(Keys.RIGHT).perform()
                    ActionBuilder(driver).clear_actions()
                    time.sleep(3)
            else:
                ActionChains(driver).send_keys(Keys.RIGHT).perform()
                ActionBuilder(driver).clear_actions()
                time.sleep(2)
            for i in range(1, 21):
                record = {col: None for col in columns}
                source_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[4]/span/a/font'
                paper_name_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                publish_time_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[5]'
                quote_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[6]/span'
                download_num_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[7]/span/span/a'

                xpaths = [source_xpath, paper_name_xpath, publish_time_xpath, quote_xpath, download_num_xpath]
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future_elements = [executor.submit(get_info, driver, xpath) for xpath in xpaths]
                source, paper_name, publish_time, quote_num, download_num= [future.result() for future in
                                                                           future_elements]
                if not quote_num.isdigit():
                    quote_num = '0'
                if not download_num.isdigit():
                    download_num = '0'
                #print(f"{source} {paper_name} {author} {publish_time} {quote_num} {download_num} \n")

                record["期刊名称"] = source
                record["论文标题"] = paper_name
                record["发表时间"] = publish_time
                record["引用量"] = quote_num
                record["下载量"] = download_num
                record["正文"] = f"{k}_{i}"

                search = '/html/body/div[2]/div[2]/div[2]/div[2]/div/div[2]/div/div[1]/div/div/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                search_click = driver.find_element(By.XPATH, search)
                driver.execute_script("arguments[0].click();", search_click)
                #time.sleep(1)
                # 获取所有窗口
                windows = driver.window_handles
                # 切换到论文详情页面对应的窗口
                #print(driver.title)
                driver.switch_to.window(windows[-1])
                #print(driver.title)
                if driver.title == '知网节':
                    record["单位"] = '未能成功加载页面'
                    record["摘要"] = '未能成功加载页面'
                    record["关键词"] = '未能成功加载页面'
                    record["分类号"] = '未能成功加载页面'
                    record["页数"] = '未能成功加载页面'
                    record["专题"] = '未能成功加载页面'
                    record["专辑"] = '未能成功加载页面'
                    record["作者姓名"] = '未能成功加载页面'

                    # time.sleep(1)
                    driver.close()
                    # time.sleep(1)
                    # 切换到检索目录
                    driver.switch_to.window(windows_init[0])
                    if os.path.exists(file_path):
                        # 文件存在：加载并追加
                        wb = load_workbook(file_path)
                        ws = wb.active
                    else:
                        # 文件不存在：创建并写入表头和第一行
                        wb = Workbook()
                        ws = wb.active
                        ws.append(columns)
                    # 保证列顺序匹配 headers
                    row = [record.get(col, "") for col in columns]
                    ws.append(row)
                    wb.save(file_path)
                    # 打印
                    print(record)
                    continue
                
                # 通过循环滚动页面来确保所有动态内容加载完成
                last_height = 0
                for _ in range(3):  # 尝试滚动3次
                    # 滚动到页面底部
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(2)  # 等待内容加载
                    
                    # 获取当前滚动高度
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    
                    # 如果高度没有变化，说明已经到底部，不再需要滚动
                    if new_height == last_height:
                        break
                    
                    last_height = new_height
                
                # 最后再滚动回中部位置，确保视野在页面中间
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
                time.sleep(1)
                
                # 定义清洗参考文献的函数
                def clean_reference(ref_text):
                    # 去除引用编号 [数字]
                    cleaned = re.sub(r'^\s*\[\d+\]\s*', '', ref_text)
                    # 只保留第一句话（通常是标题）- 在.或[J]或[C]等标记处分割
                    parts = re.split(r'(?:\.\s|\[J\]|\[C\]|\[M\])', cleaned, 1)
                    if parts:
                        title = parts[0].strip()
                        # 去除可能存在的其他括号和内部内容，如 [5] 这样的引用编号
                        title = re.sub(r'\[[^\]]*\]', '', title).strip()
                        return title
                    return cleaned
                
                #workplace_xpath='/html/body/div[2]/div[1]/div[3]/div/div[3]/div[1]/div/h3[2]/span/a'
                # abstract_xpath= '//*[@id="ChDivSummary"]/text()'
                # keyword_xpath='/html/body/div[2]/div[1]/div[3]/div/div[3]/div[3]/p'
                # Classify_number_xpath ='/html/body/div[2]/div[1]/div[3]/div/div[3]/div[4]/ul/li[3]/p'
                pages_xpath='//*[@id="DownLoadParts"]/div/div/div/p/span[3]'
                #print('正在获取workplace...')

                try:
                    # 查找所有 class="author" 的 h3 标签
                    all_author_h3s = WebDriverWait(driver, 3).until(
                        EC.presence_of_all_elements_located((By.CLASS_NAME, "author"))
                    )

                    # 排除 id="authorpart" 的那个（也就是作者名部分）
                    org_h3 = None
                    for h3 in all_author_h3s:
                        if h3.get_attribute("id") != "authorpart":
                            org_h3 = h3
                            break

                    if org_h3 is None:
                        raise Exception("未找到单位信息区域")

                    # 获取该 <h3> 下的所有 <a> 和 <span>
                    org_elements = org_h3.find_elements(By.TAG_NAME, "span")

                    orgs = []
                    for elem in org_elements:
                        text = elem.text.strip()
                        orgs.append(text)

                    workplace = "；".join(orgs) if orgs else "无"

                except:
                    workplace = "无"

                #print('正在获取abstract...')
                try:
                    abstract = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "abstract-text"))).text
                except:
                    abstract = '无'

                #print(abstract + '\n')

                # print('正在获取author...')
                try:
                    author_div = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "author"))
                    )
                    author_links = author_div.find_elements(By.TAG_NAME, "span")

                    authors = []
                    for a in author_links:
                        text = a.text.strip()
                        authors.append(text)

                    author = "；".join(authors) if authors else "无"
                except:
                    author = "无"
                # print(author + '\n')
                # print('正在获取references...')
                try:
                    references_div = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.ID, "quoted-references"))
                    )
                    references_links = references_div.find_elements(By.CSS_SELECTOR, ".ebBd > li")

                    references = []
                    for r in references_links:
                        try:
                            # 获取所有带title属性的a标签
                            title_links = r.find_elements(By.CSS_SELECTOR, "a[title]")
                            if title_links:
                                # 获取链接的title属性值作为论文标题
                                paper_titles = [link.get_attribute("title").strip() for link in title_links if link.get_attribute("title")]
                                # 只添加非空标题
                                if paper_titles:
                                    # 清洗标题，移除HTML标签和特殊格式
                                    title = paper_titles[0]
                                    # 清除类似 <Emphasis Type="SmallCaps">Pascal</Emphasis> 这样的标记
                                    title = re.sub(r'<[^>]+>([^<]+)</[^>]+>', r'\1', title)
                                    # 移除其他可能的HTML标签
                                    title = re.sub(r'<[^>]*>', '', title)
                                    cleaned_text = clean_reference(title)
                                    references.append(cleaned_text)
                            else:
                                # 如果没有找到带title的链接，则获取整个li的文本
                                r_text = r.text.strip()
                                if r_text:
                                    # 清洗获取到的文本
                                    cleaned_text = clean_reference(r_text)
                                    references.append(cleaned_text)
                        except Exception as e:
                            print(f"提取引用标题时出错: {e}")
                            # 退回到获取完整文本
                            r_text = r.text.strip()
                            if r_text:
                                # 清洗获取到的文本
                                cleaned_text = clean_reference(r_text)
                                references.append(cleaned_text)

                    reference = "；".join(references) if references else "无"
                except:
                    reference = "无"
                # print(references + '\n')

                #print('正在获取keywords...')
                try:
                    keywords = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "keywords"))).text[:-1]
                except:
                    keywords = '无'
                #print(keywords + '\n')
                #print('正在获取funds...')
                try:
                    funds = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "funds"))).text[:-1]
                except:
                    funds = '无'
                #print(funds + '\n')

                #print('正在获取publication、topic、Classify_number...')
                rowtit_elements = driver.find_elements(By.CLASS_NAME, "rowtit")
                for span in rowtit_elements:
                    text = span.text.strip()
                    if "专辑" in text:
                        # 找到父级 li 元素，再找其中的 <p>
                        parent_li = span.find_element(By.XPATH, "./ancestor::li")
                        p_elem = parent_li.find_element(By.TAG_NAME, "p")
                        #print("专辑内容:", p_elem.text.strip())
                        topic = p_elem.text.strip()
                    if "专题" in text:
                        # 找到父级 li 元素，再找其中的 <p>
                        parent_li = span.find_element(By.XPATH, "./ancestor::li")
                        p_elem = parent_li.find_element(By.TAG_NAME, "p")
                        #print("专题内容:", p_elem.text.strip())
                        publication = p_elem.text.strip()
                    if "分类号" in text:
                        # 找到父级 li 元素，再找其中的 <p>
                        parent_li = span.find_element(By.XPATH, "./ancestor::li")
                        p_elem = parent_li.find_element(By.TAG_NAME, "p")
                        #print("分类号内容:", p_elem.text.strip())
                        Classify_number = p_elem.text.strip()


                #print('正在获取pages...')

                try:
                    pages = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                        (By.XPATH, pages_xpath))).text
                except:
                    pages = '无'

                #print(pages + '\n')

                # xpaths = [workplace_xpath, abstract_xpath, keyword_xpath, Classify_number_xpath, pages_xpath]
                # with concurrent.futures.ThreadPoolExecutor() as executor:
                #     future_elements = [executor.submit(get_info, driver, xpath) for xpath in xpaths]
                # workplace, abstract, keyword, Classify_number, pages= [future.result() for future in
                #                                                                      future_elements]
                # if not pages.isdigit():
                #     pages = '0'
                # print(f"{workplace} {abstract} {keyword} {Classify_number} {pages}\n")
                record["单位"] = workplace
                record["摘要"] = abstract
                record["关键词"] = keywords
                record["基金资助"] = funds
                record["分类号"] = Classify_number
                record["页数"] = pages
                record["专题"] = publication
                record["专辑"] = topic
                record["作者姓名"] = author
                record["参考文献"] = reference


                #time.sleep(1)
                driver.close()
                #time.sleep(1)
                # 切换到检索目录
                driver.switch_to.window(list_handle)
                # df = pd.DataFrame([record])
                # df.to_excel("论文信息表.xlsx", index=False)
                if os.path.exists(file_path):
                    # 文件存在：加载并追加
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    # 文件不存在：创建并写入表头和第一行
                    wb = Workbook()
                    ws = wb.active
                    ws.append(columns)
                # 保证列顺序匹配 headers
                row = [record.get(col, "") for col in columns]
                ws.append(row)
                wb.save(file_path)
                # 打印
                print(record)
            #time.sleep(1)
    except Exception as e:
        print(e)
    time.sleep(1)
    driver.close()

if __name__ == '__main__':
    kw = input("请输入要搜索的关键词：")
    start = int(input("请输入要下载的起始页数："))
    pg = int(input("请输入要下载的页数："))
    time.sleep(2)
    try:
        crawl(start,kw, pg)
    except Exception as e:
        print(e)




