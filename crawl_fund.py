import xlwt
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
import os
from idlelib.iomenu import encoding
import concurrent.futures
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionBuilder
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import subprocess
from selenium.webdriver.common.keys import Keys
from urllib.request import urlretrieve
import random
import requests
import json
import io
import base64
import numpy as np
from PIL import Image
import re
file_path = "论文信息表_1_补.xlsx"
# 定义列标题
columns = [
    "论文标题","基金资助","发表时间"
]


def click_cancle(dr,xpath):
    try:
        # 点击取消
        cancel = dr.find_element(By.XPATH, xpath)
        dr.execute_script("arguments[0].click();", cancel)

    except Exception as e:
        print(e)

def cancle(driver):
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[2]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[3]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[4]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[7]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[8]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[9]/i')
    click_cancle(driver,'/html/body/div[2]/div[2]/div[2]/div[3]/ul[1]/li[10]/i')


from selenium import webdriver


def create_edge_driver(*, headless=False):
    # 配置浏览器选项
    options = webdriver.EdgeOptions()

    # 启用无头模式（可选）
    if headless:
        options.add_argument('--headless')


    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    # 初始化浏览器驱动
    browser = webdriver.Edge(options=options)
    #print(browser.title)

    return browser

def get_info(driver, xpath):
    try:
        element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return element.text
    except:
        return '无'

def crawl(start,key, page):
    # browser_path = "C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    # subprocess.Popen([browser_path, '--remote-debugging-port=9222'])
    driver = create_edge_driver(headless=False)
    time.sleep(1)
    driver.get("https://www.cnki.net/")  # 打开网页
    #driver.maximize_window()                      #最大化窗口
    time.sleep(1)  # 加载等待
    # txt_SearchText 为搜索框的位置  键入搜索内容
    driver.find_element(By.ID,'txt_SearchText').send_keys(key)
    # 只筛选期刊
    cancle(driver)
    # 定位到搜索按钮的位置并点击
    action = ActionChains(driver)
    action.move_to_element(driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[1]/div/div[1]/i')).perform()
    search = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[1]/div/div[2]/ul[3]/li[2]/a')
    driver.execute_script("arguments[0].click();", search)
    time.sleep(2)
    ActionBuilder(driver).clear_actions()
    # 定位到搜索按钮的位置并点击
    search = driver.find_element(By.XPATH,'/html/body/div[2]/div[2]/div[2]/div[1]/input[2]')
    driver.execute_script("arguments[0].click();", search)
    time.sleep(2) #等待加载
    # 获取初始所有窗口
    windows_init = driver.window_handles
    # 切换到目录页面对应的窗口
    driver.switch_to.window(windows_init[0])
    list_handle = driver.current_window_handle
    print(driver.title)
    try:
        for k in range(start,page):
            # page = 'page' + str(k)
            # driver.find_element(By.ID,page).click()
            if k == start:
                for t in range(1, start):
                    ActionChains(driver).send_keys(Keys.RIGHT).perform()
                    ActionBuilder(driver).clear_actions()
                    time.sleep(3)
            else:
                ActionChains(driver).send_keys(Keys.RIGHT).perform()
                ActionBuilder(driver).clear_actions()
                time.sleep(2)
            for i in range(1, 21):
                record = {col: None for col in columns}
                paper_name_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                publish_time_xpath = '//*[@id="gridTable"]/div/div/div/table/tbody/tr[' + str(i) + ']/td[5]'

                xpaths = [paper_name_xpath, publish_time_xpath]
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future_elements = [executor.submit(get_info, driver, xpath) for xpath in xpaths]
                paper_name, publish_time= [future.result() for future in future_elements]

                #print(f"{source} {paper_name} {author} {publish_time} {quote_num} {download_num} \n")


                record["论文标题"] = paper_name
                record["发表时间"] = publish_time
                record["正文"] = f"{k}_{i}"

                search = '/html/body/div[2]/div[2]/div[2]/div[2]/div/div[2]/div/div[1]/div/div/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                search_click = driver.find_element(By.XPATH, search)
                driver.execute_script("arguments[0].click();", search_click)
                #time.sleep(1)
                # 获取所有窗口
                windows = driver.window_handles
                # 切换到论文详情页面对应的窗口
                #print(driver.title)
                driver.switch_to.window(windows[-1])
                #print(driver.title)
                if driver.title == '知网节':
                    record["基金资助"] = '未能成功加载页面'

                    # time.sleep(1)
                    driver.close()
                    # time.sleep(1)
                    # 切换到检索目录
                    driver.switch_to.window(windows_init[0])
                    if os.path.exists(file_path):
                        # 文件存在：加载并追加
                        wb = load_workbook(file_path)
                        ws = wb.active
                    else:
                        # 文件不存在：创建并写入表头和第一行
                        wb = Workbook()
                        ws = wb.active
                        ws.append(columns)
                    # 保证列顺序匹配 headers
                    row = [record.get(col, "") for col in columns]
                    ws.append(row)
                    wb.save(file_path)
                    # 打印
                    print(record)
                    continue
                #print('正在获取funds...')
                try:
                    funds = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "funds"))).text[:-1]
                except:
                    funds = '无'
                #print(funds + '\n')
             
                record["基金资助"] = funds


                #time.sleep(1)
                driver.close()
                #time.sleep(1)
                # 切换到检索目录
                driver.switch_to.window(list_handle)
                # df = pd.DataFrame([record])
                # df.to_excel("论文信息表.xlsx", index=False)
                if os.path.exists(file_path):
                    # 文件存在：加载并追加
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    # 文件不存在：创建并写入表头和第一行
                    wb = Workbook()
                    ws = wb.active
                    ws.append(columns)
                # 保证列顺序匹配 headers
                row = [record.get(col, "") for col in columns]
                ws.append(row)
                wb.save(file_path)
                # 打印
                print(record)
            #time.sleep(1)
    except Exception as e:
        print(e)
    time.sleep(1)
    driver.close()

if __name__ == '__main__':
    kw = input("请输入要搜索的关键词：")
    start = int(input("请输入要下载的起始页数："))
    pg = int(input("请输入要下载的页数："))
    time.sleep(2)
    try:
        crawl(start,kw, pg)
    except Exception as e:
        print(e)




